import os
import openpyxl
from UIDBOT.parameters import Parameters
from copy import copy


def filterData(sheet, filterBy: str = "blank"):
    """"""
    rows = list(sheet.iter_rows(min_row=1, max_row=sheet.max_row))
    for row in reversed(rows):
        if row[0].row == 1:
            break
        if row[5].value != filterBy:
            sheet.delete_rows(row[0].row, 1)


def delUnusedColumns(sheet):
    sheet.delete_cols(8, 9)


class PrepareExcel:

    def __init__(self):
        self.input = os.path.join(Parameters.PATH, Parameters.fileout)
        self.workbook = openpyxl.load_workbook(self.input)
        self.inputSheet = self.workbook['Table 1']

    def copyFormat(self):
        """ copy format to column"""
        # copy format to column
        cell = self.inputSheet['A2']
        for new_cell in self.inputSheet['F']:
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)

        # copy format of header
        cell = self.inputSheet['A1']
        new_cell = self.inputSheet['F1']
        new_cell.font = copy(cell.font)
        new_cell.border = copy(cell.border)
        new_cell.fill = copy(cell.fill)

    def copyTab(self, tabname: str):
        """ copy first tab """
        newsheet = self.workbook.copy_worksheet(self.inputSheet)
        newsheet.title = tabname
        return newsheet

    def prepSheet(self, sheetname, filterrow):
        newsheet = self.copyTab(sheetname)
        filterData(newsheet, filterrow)
        delUnusedColumns(newsheet)

    def prepFile(self):
        self.copyFormat()
        self.prepSheet('blanks', 'blank')
        self.prepSheet('not valid', 'not valid')
        self.workbook.save(Parameters.fileout)


if __name__ == "__main__":
    prepXL = PrepareExcel()
    prepXL.prepFile()

