import openpyxl
from Scraper import Scraper


class ReadData:

    def __init__(self):
        self.file = r"C:\Users\wise\Documents\Projects\TAG\RPA\Long List\RPA-Prozessevaluierung Accounting v. " \
                    r"25.05.2021.xlsx"
        self.workbook = openpyxl.load_workbook(self.file)
        self.worksheet = self.workbook['Prozessevaluierung TAG']
        # self.scraper = Scraper()

    def readData(self, max_col=None):

        if max_col is None:
            max_col = self.worksheet.max_column

        colString = [""]*max_col
        for ix, col in enumerate(self.worksheet.iter_cols(max_col=max_col)):
            for cell in col:
                cellString = str(cell.value).replace("?", "").strip()
                colString[ix] = colString[ix] + cellString + "|"

        return colString

    def saveData(self, name = 'temp'):
        self.workbook.save(filename=name + '.xlsx')




rd = ReadData()
temp = rd.readData()